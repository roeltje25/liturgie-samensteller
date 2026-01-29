"""Service for exporting liturgy data to Excel registration file."""

import os
import re
import shutil
import tempfile
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formula.translate import Translator

from ..models import Liturgy, SectionType
from ..logging_config import get_logger

logger = get_logger("excel_service")


def _get_current_year() -> int:
    """Get the current year."""
    from datetime import date
    return date.today().year


def _normalize_for_match(text: str) -> str:
    """Normalize text for case-insensitive comparison."""
    if not text:
        return ""
    # Convert to lowercase, normalize whitespace
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def _fuzzy_match(query: str, text: str) -> float:
    """Calculate fuzzy match score between query and text.

    Returns a score from 0.0 (no match) to 1.0 (perfect match).
    """
    if not query or not text:
        return 0.0

    query_norm = _normalize_for_match(query)
    text_norm = _normalize_for_match(text)

    if not query_norm or not text_norm:
        return 0.0

    # Exact match
    if query_norm == text_norm:
        return 1.0

    # One contains the other
    if query_norm in text_norm or text_norm in query_norm:
        return 0.95

    # Check character overlap
    query_chars = set(query_norm.replace(' ', ''))
    text_chars = set(text_norm.replace(' ', ''))
    if query_chars and text_chars:
        overlap = len(query_chars & text_chars)
        total = len(query_chars | text_chars)
        return overlap / total * 0.8

    return 0.0


class ExcelService:
    """Service for exporting liturgy to Excel registration file."""

    # Sheet names
    LITURGY_SHEET = "Liturgiën"
    SONGS_SHEET = "Liederen"

    # Known column headers (lowercase for matching)
    COL_ZONDAG = "zondag"
    COL_DIENSTLEIDER = "dienstleider"
    COL_LITURGIE_BEKEND = "liturgie bekend"
    COL_LIED = "lied"
    COL_LIEDBUNDEL = "liedbundel"

    def __init__(self, excel_path: str):
        """Initialize the Excel service.

        Args:
            excel_path: Path to the LiederenRegister.xlsx file.
        """
        self.excel_path = excel_path

    def export_liturgy(self, liturgy: Liturgy) -> str:
        """Export liturgy to Excel file.

        Args:
            liturgy: The liturgy to export.

        Returns:
            Path to the updated Excel file.

        Raises:
            FileNotFoundError: If the Excel file doesn't exist.
            ValueError: If the liturgy has no service date.
        """
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        if not liturgy.service_date:
            raise ValueError("Liturgy has no service date set")

        # Parse service date
        service_date = date.fromisoformat(liturgy.service_date)

        # First, load with data_only=True to find the correct row
        # (dates may be formulas, need calculated values for comparison)
        target_row = None
        liturgy_sheet_name = None
        songs_sheet_name = None

        wb_readonly = load_workbook(self.excel_path, data_only=True)
        try:
            liturgy_sheet_name = self._find_sheet_name(wb_readonly, self.LITURGY_SHEET)
            songs_sheet_name = self._find_sheet_name(wb_readonly, self.SONGS_SHEET)

            if liturgy_sheet_name:
                ws_readonly = wb_readonly[liturgy_sheet_name]
                col_map = self._get_column_map(ws_readonly)
                date_col = col_map.get(self.COL_ZONDAG, 1)
                target_row = self._find_liturgy_row(ws_readonly, service_date, date_col)
        finally:
            wb_readonly.close()

        # Now load for editing (without data_only)
        wb = load_workbook(self.excel_path)

        try:
            # Update Liturgiën sheet
            if liturgy_sheet_name and liturgy_sheet_name in wb.sheetnames:
                ws_liturgy = wb[liturgy_sheet_name]
                self._update_liturgy_sheet_at_row(ws_liturgy, liturgy, service_date, target_row)

            # Get song titles
            song_titles = self._get_song_titles(liturgy)

            # Update Liederen sheet - add missing songs
            if songs_sheet_name and songs_sheet_name in wb.sheetnames:
                ws_songs = wb[songs_sheet_name]

                # First, ensure the year's named range exists
                current_year = _get_current_year()
                self._create_year_named_range(wb, current_year)

                # Ensure current year column exists in the table
                self._ensure_year_column_exists(ws_songs, wb)

                # Add missing songs
                if song_titles:
                    for title in song_titles:
                        self._ensure_song_exists(ws_songs, title)

            # Save workbook to temp file first, then replace original
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)

            try:
                wb.save(temp_path)
                wb.close()
                wb = None

                # Replace original with temp file
                shutil.move(temp_path, self.excel_path)
            except Exception:
                # Clean up temp file on error
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                raise

        finally:
            if wb:
                wb.close()

        return self.excel_path

    def get_dienstleiders(self) -> List[str]:
        """Get list of unique Dienstleiders for autocomplete.

        Returns:
            List of unique dienstleider names, sorted alphabetically.
        """
        if not os.path.exists(self.excel_path):
            logger.debug(f"Excel file not found for dienstleiders: {self.excel_path}")
            return []

        wb = None
        try:
            logger.debug(f"Loading Excel file for dienstleiders: {self.excel_path}")
            wb = load_workbook(self.excel_path, read_only=True, data_only=True)

            if self.LITURGY_SHEET not in wb.sheetnames:
                logger.warning(f"Sheet '{self.LITURGY_SHEET}' not found in Excel file")
                return []

            ws = wb[self.LITURGY_SHEET]
            col_map = self._get_column_map(ws)
            dienstleider_col = col_map.get(self.COL_DIENSTLEIDER)

            if not dienstleider_col:
                logger.warning(f"Dienstleider column not found in sheet")
                return []

            dienstleiders = set()
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row=row, column=dienstleider_col).value
                if value and isinstance(value, str) and value.strip():
                    dienstleiders.add(value.strip())

            logger.debug(f"Found {len(dienstleiders)} unique dienstleiders")
            return sorted(dienstleiders)

        except Exception as e:
            logger.error(f"Error reading dienstleiders from Excel: {e}", exc_info=True)
            return []
        finally:
            if wb:
                try:
                    wb.close()
                except Exception as e:
                    logger.debug(f"Error closing workbook: {e}")

    def _find_sheet_name(self, wb, target_name: str) -> Optional[str]:
        """Find sheet name case-insensitively."""
        target_lower = target_name.lower()
        for name in wb.sheetnames:
            if name.lower() == target_lower:
                return name
        # Also try without special characters
        target_simple = re.sub(r'[ëé]', 'e', target_lower)
        for name in wb.sheetnames:
            name_simple = re.sub(r'[ëé]', 'e', name.lower())
            if name_simple == target_simple:
                return name
        return None

    def _get_column_map(self, ws: Worksheet) -> Dict[str, int]:
        """Read headers from row 1 and return {header_lower: column_index} map.

        This handles dynamic year columns gracefully.
        """
        columns = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                columns[str(header).lower().strip()] = col
        return columns

    def _get_song_titles(self, liturgy: Liturgy) -> List[str]:
        """Extract song titles from liturgy."""
        titles = []
        for section in liturgy.sections:
            if section.section_type == SectionType.SONG and section.name:
                titles.append(section.name)
        return titles

    def _find_liturgy_row(
        self, ws: Worksheet, service_date: date, date_col: int
    ) -> Optional[int]:
        """Find existing row for date (using calculated values).

        Returns row number if found, None if not found.
        """
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=date_col).value
            if cell_value is None:
                continue

            cell_date = self._extract_date(cell_value)
            if cell_date == service_date:
                return row

        return None

    def _update_liturgy_sheet_at_row(
        self, ws: Worksheet, liturgy: Liturgy, service_date: date, target_row: Optional[int]
    ) -> None:
        """Update liturgy data at the specified row, or append if row is None."""
        col_map = self._get_column_map(ws)

        # Get column indices
        date_col = col_map.get(self.COL_ZONDAG, 1)
        dienstleider_col = col_map.get(self.COL_DIENSTLEIDER)
        bekend_col = col_map.get(self.COL_LITURGIE_BEKEND)

        # Find Lied1-Lied12 columns
        lied_cols = []
        for i in range(1, 13):
            key = f"lied{i}"
            if key in col_map:
                lied_cols.append(col_map[key])

        # Use target row or append at end
        if target_row is None:
            # Find last row with data and append after it
            target_row = 2
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=date_col).value is not None:
                    target_row = row + 1
            # Set the date for new row
            ws.cell(row=target_row, column=date_col).value = service_date

        # Set Dienstleider
        if dienstleider_col and liturgy.dienstleider:
            ws.cell(row=target_row, column=dienstleider_col).value = liturgy.dienstleider

        # Set Liturgie bekend = "ja"
        if bekend_col:
            ws.cell(row=target_row, column=bekend_col).value = "ja"

        # Set song titles
        song_titles = self._get_song_titles(liturgy)
        for i, lied_col in enumerate(lied_cols):
            if i < len(song_titles):
                ws.cell(row=target_row, column=lied_col).value = song_titles[i]
            else:
                # Clear any existing value beyond our song count
                ws.cell(row=target_row, column=lied_col).value = None

    def _extract_date(self, value) -> Optional[date]:
        """Extract date from cell value (may be datetime, date, float, or string)."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            # Excel serial date number
            try:
                # Excel's epoch is 1899-12-30 (with the 1900 leap year bug)
                from datetime import timedelta
                excel_epoch = date(1899, 12, 30)
                return excel_epoch + timedelta(days=int(value))
            except (ValueError, OverflowError):
                return None
        if isinstance(value, str):
            # Try to parse common date formats
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def _ensure_song_exists(self, ws: Worksheet, song_title: str) -> bool:
        """Add song to Liederen sheet if not exists.

        Properly handles Excel Tables by expanding the table range.

        Returns:
            True if song was added, False if it already existed.
        """
        col_map = self._get_column_map(ws)

        # Try to find the song name column - could be "lied", "titel", "naam", etc.
        lied_col = None
        for key in [self.COL_LIED, "titel", "naam", "song", "name"]:
            if key in col_map:
                lied_col = col_map[key]
                break

        # Default to column 1 if no header found
        if lied_col is None:
            lied_col = 1

        # Check if song already exists (case-insensitive)
        song_title_clean = song_title.strip()
        song_title_lower = song_title_clean.lower()

        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=lied_col).value
            if name and isinstance(name, str):
                if name.strip().lower() == song_title_lower:
                    return False  # Already exists (exact match)

        # Song not found - add it
        # Find the table and its current range
        table = None
        table_name = None
        for tbl_name in ws.tables:
            table = ws.tables[tbl_name]
            table_name = tbl_name
            break  # Use first table

        if table:
            # Parse the table range (e.g., "A1:J359")
            from openpyxl.utils import get_column_letter
            ref_parts = table.ref.split(':')
            if len(ref_parts) == 2:
                # Get the end row and column from the table range
                import re
                end_match = re.match(r'([A-Z]+)(\d+)', ref_parts[1])
                if end_match:
                    end_col_letter = end_match.group(1)
                    end_row = int(end_match.group(2))

                    # New row is at the end of the table
                    new_row = end_row + 1

                    # Set the song name
                    ws.cell(row=new_row, column=lied_col).value = song_title_clean

                    # Copy formulas from the row above
                    self._copy_table_row_formulas(ws, end_row, new_row, lied_col)

                    # Expand the table range
                    new_ref = f"{ref_parts[0]}:{end_col_letter}{new_row}"
                    table.ref = new_ref

                    # Also update the autoFilter range if present
                    if table.autoFilter:
                        table.autoFilter.ref = new_ref

                    return True

        # Fallback: no table found, just append
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=lied_col).value = song_title_clean
        return True

    def _copy_table_row_formulas(
        self, ws: Worksheet, source_row: int, target_row: int, skip_col: int
    ) -> None:
        """Copy formulas from one row to another within a table.

        Args:
            ws: Worksheet.
            source_row: Row to copy formulas from.
            target_row: Row to copy formulas to.
            skip_col: Column to skip (song name column).
        """
        from openpyxl.formula.translate import Translator

        for col in range(1, ws.max_column + 1):
            if col == skip_col:
                continue

            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)

            # Check if source cell has a formula
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                try:
                    formula = source_cell.value
                    translated = Translator(
                        formula, origin=source_cell.coordinate
                    ).translate_formula(target_cell.coordinate)
                    target_cell.value = translated
                except Exception:
                    # If formula translation fails, skip this cell
                    pass

    def _copy_row_formulas(
        self, ws: Worksheet, source_row: int, target_row: int, skip_col: int
    ) -> None:
        """Copy formulas from one row to another, translating references.

        Args:
            ws: Worksheet.
            source_row: Row to copy formulas from.
            target_row: Row to copy formulas to.
            skip_col: Column to skip (song name column).
        """
        for col in range(1, ws.max_column + 1):
            if col == skip_col:
                continue

            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)

            # Check if source cell has a formula
            if source_cell.data_type == 'f' or (
                source_cell.value and
                isinstance(source_cell.value, str) and
                source_cell.value.startswith('=')
            ):
                try:
                    formula = source_cell.value
                    if formula:
                        translated = Translator(
                            formula, origin=source_cell.coordinate
                        ).translate_formula(target_cell.coordinate)
                        target_cell.value = translated
                except Exception:
                    # If formula translation fails, skip this cell
                    pass

    def _create_year_named_range(self, wb, year: int) -> bool:
        """Create a named range for the year if it doesn't exist.

        The named ranges point to row ranges in the Liturgiën sheet, e.g.:
        Year2025: Liturgiën!$A$319:$P$370

        Args:
            wb: The workbook.
            year: The year to create the named range for.

        Returns:
            True if created, False if already exists or couldn't be created.
        """
        from openpyxl.workbook.defined_name import DefinedName

        range_name = f"Year{year}"

        # Check if it already exists
        if range_name in wb.defined_names:
            return False

        # Find the previous year's range to determine where to start
        prev_year = year - 1
        prev_range_name = f"Year{prev_year}"

        if prev_range_name not in wb.defined_names:
            # Can't create without knowing where to start
            return False

        prev_defn = wb.defined_names[prev_range_name]
        prev_range = prev_defn.attr_text

        # Parse the previous range, e.g. "Liturgiën!$A$319:$P$370"
        match = re.match(r"(.+)!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", prev_range)
        if not match:
            return False

        sheet_name = match.group(1)
        start_col = match.group(2)
        end_col = match.group(4)
        prev_end_row = int(match.group(5))

        # Calculate new range (starting after previous year ends)
        new_start_row = prev_end_row + 1
        # Each year typically has 52-53 rows (weeks)
        new_end_row = new_start_row + 52

        new_range = f"{sheet_name}!${start_col}${new_start_row}:${end_col}${new_end_row}"

        # Create the named range
        new_defn = DefinedName(range_name, attr_text=new_range)
        wb.defined_names[range_name] = new_defn

        return True

    def _ensure_year_column_exists(self, ws: Worksheet, wb) -> bool:
        """Ensure the current year's column exists in Liederen sheet.

        The Liederen sheet has columns like "Totaal 2019", "Totaal 2020", etc.
        The last column is always "Te rapporteren".
        If the current year's column is missing, insert it before "Te rapporteren".

        This method properly handles Excel Tables by:
        1. Inserting the worksheet column
        2. Creating a TableColumn with calculatedColumnFormula
        3. Updating the table's tableColumns list
        4. Updating the table reference

        Args:
            ws: The Liederen worksheet.

        Returns:
            True if a column was added, False if it already existed.
        """
        from openpyxl.utils import get_column_letter, column_index_from_string
        from openpyxl.worksheet.table import TableColumn, TableFormula

        current_year = _get_current_year()
        year_col_name = f"Totaal {current_year}"

        # Find existing columns
        col_map = self._get_column_map(ws)

        # Check if year column already exists
        if year_col_name.lower() in col_map:
            return False

        # Find "Te rapporteren" column (should be the last column)
        te_rapporteren_col = None
        for header, col_idx in col_map.items():
            if "te rapporteren" in header.lower():
                te_rapporteren_col = col_idx
                break

        if te_rapporteren_col is None:
            # No "Te rapporteren" column found, skip
            return False

        # Find the table
        table = None
        for tbl_name in ws.tables:
            table = ws.tables[tbl_name]
            break

        if not table:
            # No table, skip year column creation
            return False

        # Parse table reference to get table start column
        ref_parts = table.ref.split(':')
        if len(ref_parts) != 2:
            return False

        start_match = re.match(r'([A-Z]+)(\d+)', ref_parts[0])
        end_match = re.match(r'([A-Z]+)(\d+)', ref_parts[1])
        if not start_match or not end_match:
            return False

        table_start_col = column_index_from_string(start_match.group(1))
        end_col_num = column_index_from_string(end_match.group(1))
        end_row = end_match.group(2)

        # Calculate the index in the table columns (0-based, relative to table start)
        insert_table_idx = te_rapporteren_col - table_start_col

        # Insert the worksheet column
        ws.insert_cols(te_rapporteren_col)

        # Set the header
        ws.cell(row=1, column=te_rapporteren_col).value = year_col_name

        # Create the TableColumn with proper formula
        year_named_range = f"Year{current_year}"
        formula_text = f'COUNTIF({year_named_range},Table1[[#This Row],[Lied]])'
        calc_formula = TableFormula(attr_text=formula_text)

        # Get max id from existing columns
        max_id = max(tc.id for tc in table.tableColumns)

        new_table_col = TableColumn(
            id=max_id + 1,
            name=year_col_name,
            calculatedColumnFormula=calc_formula
        )

        # Insert the new column into tableColumns at the correct position
        cols_list = list(table.tableColumns)
        cols_list.insert(insert_table_idx, new_table_col)
        table.tableColumns = cols_list

        # Update the table reference (expand by one column)
        new_end_col_letter = get_column_letter(end_col_num + 1)
        new_ref = f"{ref_parts[0]}:{new_end_col_letter}{end_row}"
        table.ref = new_ref

        # Update autoFilter if present
        if table.autoFilter:
            table.autoFilter.ref = new_ref

        # Set cell formulas explicitly using the new year's named range
        # The formula counts how many times the song appears in the year's range
        # Use full structured reference syntax: Table1[[#This Row],[Lied]]
        for row in range(2, ws.max_row + 1):
            target_cell = ws.cell(row=row, column=te_rapporteren_col)
            target_cell.value = f'=COUNTIF({year_named_range},Table1[[#This Row],[Lied]])'

        return True
